import openpyxl
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('docs/CarbonCalculator_Examples_V3.0_August2025.xlsx', data_only=True)
ws = wb['Natural regeneration']

# Check dimensions
print(f"Sheet dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Find all non-empty cells beyond row 51
print("\n=== ALL NON-EMPTY CELLS rows 51+ ===")
for row in ws.iter_rows(min_row=51, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None:
            print(f'{cell.coordinate}: {repr(cell.value)}')

# Also check the area after column AZ in early rows
print("\n=== COLUMNS AZ+ in rows 1-30 ===")
for row in ws.iter_rows(min_row=1, max_row=30, min_col=52, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None:
            print(f'{cell.coordinate}: {repr(cell.value)}')
