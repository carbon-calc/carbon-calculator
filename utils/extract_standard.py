"""
Extract ONLY the StandardProjectCarbonCalculator sheet - with FULL formula detail.
"""
import openpyxl
from openpyxl.utils import get_column_letter

XLSX = 'docs/CarbonCalculator_V3.0_August2025.xlsx'
wb_f = openpyxl.load_workbook(XLSX, data_only=False)
wb_d = openpyxl.load_workbook(XLSX, data_only=True)

ws_f = wb_f['StandardProjectCarbonCalculator']
ws_d = wb_d['StandardProjectCarbonCalculator']

print(f"Sheet: {ws_f.max_row} rows x {ws_f.max_column} cols")

for r in range(1, ws_f.max_row + 1):
    entries = []
    for c in range(1, ws_f.max_column + 1):
        vf = ws_f.cell(row=r, column=c).value
        vd = ws_d.cell(row=r, column=c).value
        cl = get_column_letter(c)
        if vf is not None:
            if isinstance(vf, str) and vf.startswith('='):
                entries.append(f"{cl}{r}=[F]{vf} => {repr(vd)}")
            else:
                entries.append(f"{cl}{r}={repr(vf)}")
    if entries:
        for e in entries:
            print(f"  {e}")
        print()

wb_f.close()
wb_d.close()
