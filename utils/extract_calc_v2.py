"""
Extract detailed calculation logic from the key computation sheets.
"""
import openpyxl
from openpyxl.utils import get_column_letter

XLSX = 'docs/CarbonCalculator_V3.0_August2025.xlsx'
wb_f = openpyxl.load_workbook(XLSX, data_only=False)
wb_d = openpyxl.load_workbook(XLSX, data_only=True)

def dump_sheet(name, max_rows=300, max_cols=40):
    if name not in wb_f.sheetnames:
        print(f"\n!!! Sheet '{name}' NOT FOUND !!!")
        return
    ws_f = wb_f[name]
    ws_d = wb_d[name]
    print(f"\n{'='*80}")
    print(f"SHEET: {name} ({ws_f.max_row} rows x {ws_f.max_column} cols)")
    print(f"{'='*80}")
    for r in range(1, min(ws_f.max_row + 1, max_rows + 1)):
        entries = []
        for c in range(1, min(ws_f.max_column + 1, max_cols + 1)):
            vf = ws_f.cell(row=r, column=c).value
            vd = ws_d.cell(row=r, column=c).value
            cl = get_column_letter(c)
            if vf is not None:
                if isinstance(vf, str) and vf.startswith('='):
                    entries.append(f"{cl}{r}=[F]{vf[:120]} => {repr(vd)[:50]}")
                else:
                    entries.append(f"{cl}{r}={repr(vf)[:80]}")
        if entries:
            print(f"  {' | '.join(entries)}")

# Key computation sheets
for s in ['Carbon sequestration table', 'PIU vintage', 'Establishment emissions',
          'Soil carbon', 'Planting details', 'Project details']:
    dump_sheet(s)

# Also check for biomass lookup structure
dump_sheet('Biomass carbon lookup table', max_rows=15, max_cols=15)

# Check all sheet names
print("\n=== ALL SHEET NAMES ===")
for name in wb_f.sheetnames:
    ws = wb_f[name]
    print(f"  {name}: {ws.max_row}r x {ws.max_column}c")

# Named ranges
print("\n=== DEFINED NAMES ===")
for name, dn in wb_f.defined_names.items():
    try:
        dests = list(dn.destinations)
        print(f"  {name}: {dests}")
    except:
        print(f"  {name}: (could not resolve)")

wb_f.close()
wb_d.close()
