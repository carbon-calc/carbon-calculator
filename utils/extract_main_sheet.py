"""
Extract the StandardProjectCarbonCalculator sheet completely - formulas + data.
"""
import openpyxl
from openpyxl.utils import get_column_letter

XLSX = 'docs/CarbonCalculator_V3.0_August2025.xlsx'
wb_f = openpyxl.load_workbook(XLSX, data_only=False)
wb_d = openpyxl.load_workbook(XLSX, data_only=True)

ws_f = wb_f['StandardProjectCarbonCalculator']
ws_d = wb_d['StandardProjectCarbonCalculator']

print(f"StandardProjectCarbonCalculator: {ws_f.max_row} rows x {ws_f.max_column} cols")
print()

for r in range(1, ws_f.max_row + 1):
    entries = []
    for c in range(1, ws_f.max_column + 1):
        vf = ws_f.cell(row=r, column=c).value
        vd = ws_d.cell(row=r, column=c).value
        cl = get_column_letter(c)
        if vf is not None:
            if isinstance(vf, str) and vf.startswith('='):
                entries.append(f"{cl}{r}=[F]{vf[:200]} => {repr(vd)[:60]}")
            else:
                entries.append(f"{cl}{r}={repr(vf)[:100]}")
    if entries:
        print(f"  {' | '.join(entries)}")

# Also dump clearfell max seq structure
print("\n=== Clearfell max seq values (first 30 rows) ===")
ws2f = wb_f['Clearfell max seq values']
ws2d = wb_d['Clearfell max seq values']
print(f"Size: {ws2f.max_row} rows x {ws2f.max_column} cols")
for r in range(1, 31):
    entries = []
    for c in range(1, min(ws2f.max_column + 1, 38)):
        vf = ws2f.cell(row=r, column=c).value
        vd = ws2d.cell(row=r, column=c).value
        cl = get_column_letter(c)
        if vf is not None:
            if isinstance(vf, str) and vf.startswith('='):
                entries.append(f"{cl}{r}=[F]{vf[:150]} => {repr(vd)[:40]}")
            else:
                entries.append(f"{cl}{r}={repr(vf)[:80]}")
    if entries:
        print(f"  {' | '.join(entries)}")

wb_f.close()
wb_d.close()
