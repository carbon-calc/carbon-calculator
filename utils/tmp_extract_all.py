import openpyxl
import warnings
import json
from datetime import datetime

warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('docs/CarbonCalculator_Examples_V3.0_August2025.xlsx', data_only=True)

EXAMPLE_SHEETS = [
    'Natural regeneration',
    'Broadleaved min intervention',
    'Mixed conifer thin',
    'Conifer clearfell',
    'Small Mixed Wood',
]

def safe_val(v):
    """Convert cell values to JSON-serializable types."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    return v

def extract_sheet(ws, sheet_name):
    data = {"sheetName": sheet_name}
    
    # === PROJECT CONFIGURATION ===
    data["projectConfig"] = {
        "projectName": safe_val(ws['B5'].value),
        "calculatedBy": safe_val(ws['B6'].value),
        "dateCompleted": safe_val(ws['B7'].value),
        "calculatorVersion": safe_val(ws['B8'].value),
        "projectStartDate": safe_val(ws['E10'].value),
        "projectDuration": safe_val(ws['E11'].value),
        "netArea": safe_val(ws['E12'].value),
        "country": safe_val(ws['E13'].value),
        "woodlandCarbonGuarantee": safe_val(ws['E14'].value),
        "verificationFrequency": safe_val(ws['E15'].value),
    }
    
    # === PLANTING SECTIONS (rows 6-30, columns G-W) ===
    sections = []
    for row_num in range(6, 31):
        species = ws.cell(row=row_num, column=8).value  # H - Species
        if species and 'Select' not in str(species):
            section = {
                "row": row_num,
                "sectionNumber": safe_val(ws.cell(row=row_num, column=7).value),  # G
                "species": safe_val(species),  # H
                "plannedSpacing": safe_val(ws.cell(row=row_num, column=9).value),  # I
                "speciesModel": safe_val(ws.cell(row=row_num, column=10).value),  # J
                "lookupSpacing": safe_val(ws.cell(row=row_num, column=12).value),  # L
                "yieldClassESC": safe_val(ws.cell(row=row_num, column=14).value),  # N
                "yieldClassLookup": safe_val(ws.cell(row=row_num, column=16).value),  # P
                "managementRegime": safe_val(ws.cell(row=row_num, column=18).value),  # R
                "percentageOfArea": safe_val(ws.cell(row=row_num, column=19).value),  # S
                "areaHectares": safe_val(ws.cell(row=row_num, column=20).value),  # T
                "clearfellAge": safe_val(ws.cell(row=row_num, column=21).value),  # U
                "clearfellCap": safe_val(ws.cell(row=row_num, column=22).value),  # V
            }
            sections.append(section)
    
    # Also check rows 31-33 for natural regen / woody shrubs / total area
    for row_num in [31, 32, 33]:
        label = ws.cell(row=row_num, column=8).value  # H
        if label:
            section = {
                "row": row_num,
                "label": safe_val(label),
                "speciesName": safe_val(ws.cell(row=row_num, column=9).value),  # I
                "note": safe_val(ws.cell(row=row_num, column=14).value),  # N
                "percentageOfArea": safe_val(ws.cell(row=row_num, column=19).value),  # S
                "areaHectares": safe_val(ws.cell(row=row_num, column=20).value),  # T
            }
            sections.append(section)
    
    data["plantingSections"] = sections
    
    # === ESTABLISHMENT EMISSIONS (rows 16-34) ===
    emissions = []
    emission_items = [
        (17, "Seedlings"),
        (18, "Tree protection - Tree shelters (1.2m)"),
        (19, "Tree protection - Spiral guards (0.6m)"),
        (20, "Tree protection - Voleguards"),
        (21, "Fertiliser"),
        (22, "Ground preparation (fuel): Mounding"),
        (23, "Ground preparation (fuel): Scarifying"),
        (24, "Ground preparation (fuel): Ploughing"),
        (25, "Ground preparation (fuel): Subsoiling"),
        (26, "Herbicide"),
    ]
    for row_num, label in emission_items:
        spacing = safe_val(ws.cell(row=row_num, column=2).value)  # B
        area = safe_val(ws.cell(row=row_num, column=3).value)  # C
        rate = safe_val(ws.cell(row=row_num, column=4).value)  # D
        total = safe_val(ws.cell(row=row_num, column=5).value)  # E
        emissions.append({
            "type": label,
            "spacing": spacing,
            "area": area,
            "ratePerHa": rate,
            "totalTCO2e": total,
        })
    
    # Fencing
    emissions.append({
        "type": "Length of fence in metres",
        "quantity": safe_val(ws['C28'].value),
        "ratePerUnit": safe_val(ws['D28'].value),
        "totalTCO2e": safe_val(ws['E28'].value),
    })
    emissions.append({
        "type": "Number of gates",
        "quantity": safe_val(ws['C29'].value),
        "ratePerUnit": safe_val(ws['D29'].value),
        "totalTCO2e": safe_val(ws['E29'].value),
    })
    
    # Road building
    emissions.append({
        "type": "Length of road in kilometres",
        "quantity": safe_val(ws['C31'].value),
        "ratePerUnit": safe_val(ws['D31'].value),
        "totalTCO2e": safe_val(ws['E31'].value),
    })
    
    # Removal of vegetation
    emissions.append({
        "type": "Removal of trees or vegetation",
        "totalTCO2e": safe_val(ws['E33'].value),
    })
    
    data["establishmentEmissions"] = emissions
    data["totalEstablishmentEmissions"] = safe_val(ws['E34'].value)
    
    # === SOIL CARBON ACCUMULATION ===
    data["soilCarbonAccumulationArea"] = safe_val(ws['E37'].value)
    
    # === BASELINE AND LEAKAGE ===
    data["baselineSignificant"] = safe_val(ws['E39'].value)
    data["leakageSignificant"] = safe_val(ws['E40'].value)
    
    # === SOIL CARBON ASSUMPTIONS (rows 44-50) ===
    soil_rows = []
    for row_num in range(45, 51):
        landuse = ws.cell(row=row_num, column=1).value  # A
        if landuse and 'Please select' not in str(landuse):
            soil_rows.append({
                "previousLanduse": safe_val(landuse),
                "soilType": safe_val(ws.cell(row=row_num, column=2).value),  # B
                "groundPreparation": safe_val(ws.cell(row=row_num, column=3).value),  # C
                "areaHectares": safe_val(ws.cell(row=row_num, column=7).value),  # G
                "percentTopsoilCarbonLost": safe_val(ws.cell(row=row_num, column=8).value),  # H
                "soilEmissionsPerHa": safe_val(ws.cell(row=row_num, column=9).value),  # I
                "soilEmissionsTotal": safe_val(ws.cell(row=row_num, column=10).value),  # J
            })
    data["soilCarbonAssumptions"] = soil_rows
    data["soilTotalArea"] = safe_val(ws['G51'].value)
    data["soilTotalEmissions"] = safe_val(ws['J51'].value)
    
    # === RESULTS TABLE (columns CB-CM, rows 6-26) ===
    # CB=80, CC=81, CD=82, CE=83, CF=84, CG=85, CH=86, CI=87, CJ=88, CK=89, CL=90, CM=91
    results = []
    for row_num in range(6, 27):
        year = ws.cell(row=row_num, column=80).value  # CB
        if year is not None:
            result = {
                "cumulativeToYear": safe_val(year),
                "A_biomassCarbon": safe_val(ws.cell(row=row_num, column=81).value),  # CC
                "B_after20pctDeduction": safe_val(ws.cell(row=row_num, column=82).value),  # CD
                "C_establishmentEmissions": safe_val(ws.cell(row=row_num, column=83).value),  # CE
                "D_soilCarbon": safe_val(ws.cell(row=row_num, column=84).value),  # CF
                "E_grossSequestration": safe_val(ws.cell(row=row_num, column=85).value),  # CG
                "F_baseline": safe_val(ws.cell(row=row_num, column=86).value),  # CH
                "G_leakage": safe_val(ws.cell(row=row_num, column=87).value),  # CI
                "H_netCarbon": safe_val(ws.cell(row=row_num, column=88).value),  # CJ
                "I_bufferDeduction": safe_val(ws.cell(row=row_num, column=89).value),  # CK
                "J_netAfterBuffer": safe_val(ws.cell(row=row_num, column=90).value),  # CL
                "K_perHectare": safe_val(ws.cell(row=row_num, column=91).value),  # CM
            }
            results.append(result)
    data["results"] = results
    
    # === PIU VINTAGE TABLE (Version 1 - standard WCC, columns CO-CT, rows 6-17) ===
    piu_v1 = []
    for row_num in range(6, 18):
        years = ws.cell(row=row_num, column=93).value  # CO
        if years is not None and years != 'Total':
            piu_v1.append({
                "yearsFromStart": safe_val(years),
                "vintageStart": safe_val(ws.cell(row=row_num, column=94).value),  # CP
                "vintageEnd": safe_val(ws.cell(row=row_num, column=95).value),  # CQ
                "totalPIU": safe_val(ws.cell(row=row_num, column=96).value),  # CR
                "bufferContribution": safe_val(ws.cell(row=row_num, column=97).value),  # CS
                "piuToProject": safe_val(ws.cell(row=row_num, column=98).value),  # CT
            })
        elif years == 'Total':
            piu_v1.append({
                "label": "Total",
                "totalPIU": safe_val(ws.cell(row=row_num, column=96).value),
                "bufferContribution": safe_val(ws.cell(row=row_num, column=97).value),
                "piuToProject": safe_val(ws.cell(row=row_num, column=98).value),
            })
    data["piuVintageV1"] = piu_v1
    
    return data

# Extract all sheets
all_data = []
for sheet_name in EXAMPLE_SHEETS:
    ws = wb[sheet_name]
    print(f"Extracting: {sheet_name}...")
    sheet_data = extract_sheet(ws, sheet_name)
    all_data.append(sheet_data)

# Write output
with open('/tmp/carbon_calc_test_data.json', 'w') as f:
    json.dump(all_data, f, indent=2)

print("\nDone! Output written to /tmp/carbon_calc_test_data.json")
print(f"Extracted {len(all_data)} sheets")
