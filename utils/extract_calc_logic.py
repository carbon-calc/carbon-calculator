"""
Extract the complete carbon calculation logic from CarbonCalculator_V3.0_August2025.xlsx.
This reads BOTH formulas and data to understand the full calculation pipeline.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

XLSX = 'docs/CarbonCalculator_V3.0_August2025.xlsx'

# --- 1. Load with formulas (not data_only) to see calculation logic ---
wb_formulas = openpyxl.load_workbook(XLSX, data_only=False)
wb_data = openpyxl.load_workbook(XLSX, data_only=True)

print("=== ALL SHEET NAMES ===")
for name in wb_formulas.sheetnames:
    ws = wb_formulas[name]
    print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")

# --- 2. Dump the main calculation sheets with formulas ---
CALC_SHEETS = [
    'Project details',
    'Planting details',
    'Establishment emissions',
    'Soil carbon',
    'Carbon sequestration table',
    'PIU vintage',
]

for sheet_name in CALC_SHEETS:
    if sheet_name not in wb_formulas.sheetnames:
        print(f"\n!!! Sheet '{sheet_name}' NOT FOUND !!!")
        continue
    ws_f = wb_formulas[sheet_name]
    ws_d = wb_data[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name} ({ws_f.max_row} rows x {ws_f.max_column} cols)")
    print(f"{'='*80}")
    for row_num in range(1, min(ws_f.max_row + 1, 200)):
        row_entries = []
        for col in range(1, min(ws_f.max_column + 1, 40)):
            val_f = ws_f.cell(row=row_num, column=col).value
            val_d = ws_d.cell(row=row_num, column=col).value
            cl = get_column_letter(col)
            if val_f is not None:
                if isinstance(val_f, str) and val_f.startswith('='):
                    row_entries.append(f"{cl}{row_num}=[FORMULA]{val_f} => {repr(val_d)[:50]}")
                else:
                    row_entries.append(f"{cl}{row_num}={repr(val_f)[:80]}")
        if row_entries:
            print(f"  {' | '.join(row_entries)}")

# --- 3. Dump lookup table structures ---
LOOKUP_SHEETS = [
    'Biomass carbon lookup table',
    'Soil_Emissions',
    'Species lookup',
    'Species_YC_Ranges',
    'Validation lists',
    'Spacing (tree numbers)',
    'PIU_calc',
]

for sheet_name in LOOKUP_SHEETS:
    if sheet_name not in wb_formulas.sheetnames:
        print(f"\n!!! Lookup sheet '{sheet_name}' NOT FOUND !!!")
        continue
    ws_f = wb_formulas[sheet_name]
    ws_d = wb_data[sheet_name]
    print(f"\n{'='*80}")
    print(f"LOOKUP: {sheet_name} ({ws_f.max_row} rows x {ws_f.max_column} cols)")
    print(f"{'='*80}")
    # For large sheets, show first 30 rows + any formula rows
    max_show = 60 if ws_f.max_row < 100 else 30
    for row_num in range(1, min(ws_f.max_row + 1, max_show)):
        row_entries = []
        for col in range(1, min(ws_f.max_column + 1, 30)):
            val_f = ws_f.cell(row=row_num, column=col).value
            val_d = ws_d.cell(row=row_num, column=col).value
            cl = get_column_letter(col)
            if val_f is not None:
                if isinstance(val_f, str) and val_f.startswith('='):
                    row_entries.append(f"{cl}{row_num}=[F]{val_f} => {repr(val_d)[:40]}")
                else:
                    row_entries.append(f"{cl}{row_num}={repr(val_f)[:60]}")
        if row_entries:
            print(f"  {' | '.join(row_entries)}")
    if ws_f.max_row > max_show:
        print(f"  ... ({ws_f.max_row - max_show} more rows)")

# --- 4. Named ranges ---
print(f"\n{'='*80}")
print("NAMED RANGES / DEFINED NAMES")
print(f"{'='*80}")
for dn in wb_formulas.defined_names.definedName:
    dests = list(dn.destinations)
    print(f"  {dn.name}: {dests}")

wb_formulas.close()
wb_data.close()
