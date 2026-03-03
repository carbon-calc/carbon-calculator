import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('docs/CarbonCalculator_V3.0_August2025.xlsx', data_only=True)

# Validation lists sheet
print("=== VALIDATION LISTS SHEET ===")
vs = wb['Validation lists']
for row_num in range(1, vs.max_row + 1):
    row_data = []
    for col in range(1, vs.max_column + 1):
        val = vs.cell(row=row_num, column=col).value
        if val is not None:
            col_letter = get_column_letter(col)
            row_data.append(f"{col_letter}:{repr(val)[:60]}")
    if row_data:
        print(f"  Row {row_num}: {' | '.join(row_data)}")

# Species lookup sheet  
print("\n=== SPECIES LOOKUP SHEET (first 20 rows) ===")
sl = wb['Species lookup']
for row_num in range(1, min(21, sl.max_row + 1)):
    row_data = []
    for col in range(1, sl.max_column + 1):
        val = sl.cell(row=row_num, column=col).value
        if val is not None:
            col_letter = get_column_letter(col)
            row_data.append(f"{col_letter}:{repr(val)[:60]}")
    if row_data:
        print(f"  Row {row_num}: {' | '.join(row_data)}")

# Species_YC_Ranges sheet
print("\n=== SPECIES_YC_RANGES SHEET ===")
sr = wb['Species_YC_Ranges']
for row_num in range(1, sr.max_row + 1):
    row_data = []
    for col in range(1, sr.max_column + 1):
        val = sr.cell(row=row_num, column=col).value
        if val is not None:
            col_letter = get_column_letter(col)
            row_data.append(f"{col_letter}:{repr(val)[:40]}")
    if row_data:
        print(f"  Row {row_num}: {' | '.join(row_data)}")

# Biomass lookup table structure
print("\n=== BIOMASS LOOKUP TABLE (first 5 rows) ===")
bl = wb['Biomass carbon lookup table']
for row_num in range(1, 6):
    row_data = []
    for col in range(1, bl.max_column + 1):
        val = bl.cell(row=row_num, column=col).value
        if val is not None:
            col_letter = get_column_letter(col)
            row_data.append(f"{col_letter}:{repr(val)[:60]}")
    if row_data:
        print(f"  Row {row_num}: {' | '.join(row_data)}")
print(f"  Total rows: {bl.max_row}")
