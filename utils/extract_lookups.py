"""Extract all lookup tables from the WCC Carbon Calculator Excel file."""
import json
import openpyxl
import sys

wb = openpyxl.load_workbook(
    "docs/CarbonCalculator_V3.0_August2025.xlsx",
    data_only=True
)

output = {}

# ============================================================================
# 1. Species Lookup — maps species common name → species model code
# ============================================================================
def extract_species_lookup():
    """Extract species to model code mapping from Species_Lookup named range."""
    ws = wb["ValidationLists"]
    species = []
    # Species lookup is typically in columns A-D of ValidationLists
    # Let's find it by scanning for species data
    # The Species_Lookup named range maps common name → code
    # Let's look for the data - scan the sheet for patterns
    
    # First, let's check defined names
    for name, dn in wb.defined_names.items():
        if "species" in name.lower() or "Species" in name:
            print(f"  Named range: {name} = {dn.attr_text}", file=sys.stderr)
    
    return species

# ============================================================================
# 2. Get all named ranges for debugging
# ============================================================================
print("=== Named Ranges ===", file=sys.stderr)
for name, dn in wb.defined_names.items():
    print(f"  {name} = {dn.attr_text}", file=sys.stderr)

# ============================================================================
# 3. Extract ValidationLists sheet structure  
# ============================================================================
ws_vl = wb["ValidationLists"]
print(f"\n=== ValidationLists: {ws_vl.max_row} rows x {ws_vl.max_column} cols ===", file=sys.stderr)

# Print first few rows to understand structure
for row in range(1, min(30, ws_vl.max_row + 1)):
    vals = []
    for col in range(1, min(15, ws_vl.max_column + 1)):
        v = ws_vl.cell(row, col).value
        if v is not None:
            vals.append(f"[{col}]={v}")
    if vals:
        print(f"  Row {row}: {', '.join(vals)}", file=sys.stderr)

# ============================================================================
# 4. Discover the biomass lookup table sheet
# ============================================================================
print(f"\n=== Sheet Names ===", file=sys.stderr)
for name in wb.sheetnames:
    ws_tmp = wb[name]
    print(f"  {name}: {ws_tmp.max_row} rows x {ws_tmp.max_column} cols", file=sys.stderr)

# ============================================================================
# 5. Extract tree spacing emission rates from ValidationLists
# ============================================================================
# The spacing/emission rates table starts at row 1-16 of ValidationLists
# Columns: spacing, seedling rate, shelter rate, spiral guard rate, voleguard rate, fertiliser rate
print(f"\n=== Spacing Emission Rates (ValidationLists rows 1-20) ===", file=sys.stderr)
for row in range(1, 21):
    vals = []
    for col in range(1, 10):
        v = ws_vl.cell(row, col).value
        if v is not None:
            vals.append(f"[{col}]={v}")
    if vals:
        print(f"  Row {row}: {', '.join(vals)}", file=sys.stderr)

# ============================================================================
# 6. Check soil carbon lookup (further down in ValidationLists)  
# ============================================================================
print(f"\n=== Soil Carbon Lookup (ValidationLists rows 15-60) ===", file=sys.stderr)
for row in range(15, min(61, ws_vl.max_row + 1)):
    vals = []
    for col in range(1, 20):
        v = ws_vl.cell(row, col).value
        if v is not None:
            vals.append(f"[{col}]={v}")
    if vals:
        print(f"  Row {row}: {', '.join(vals)}", file=sys.stderr)

print("\nDone scanning structure.", file=sys.stderr)
