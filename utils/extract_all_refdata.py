"""Extract all missing reference/lookup data from Excel into TTL and JSON-LD files."""
import json, warnings, re
warnings.filterwarnings("ignore")
import openpyxl

wb = openpyxl.load_workbook("docs/CarbonCalculator_V3.0_August2025.xlsx", data_only=True)

# ============================================================================
# 1. TREE SPECIES (TTL) — Species lookup sheet, rows 2-137, cols A-D
# ============================================================================
ws = wb["Species lookup"]
lines = [
    '@prefix rdfs:  <http://www.w3.org/2000/01/rdf-schema#> .',
    '@prefix ref:   <https://w3id.org/wcc/reference#> .',
    '',
    '# =============================================================================',
    '# Reference Data — Tree Species',
    '# =============================================================================',
    ''
]

for row in range(2, 138):
    common = ws.cell(row, 1).value
    code = ws.cell(row, 2).value
    latin = ws.cell(row, 3).value
    model = ws.cell(row, 4).value
    if not common:
        continue
    common = str(common).strip()
    code = str(code).strip() if code else ""
    latin = str(latin).strip() if latin else ""
    model = str(model).strip() if model else ""
    # Create IRI-safe local name
    local = re.sub(r'[^A-Za-z0-9]', '', common.title())
    lines.append(f'ref:{local}')
    lines.append(f'    a ref:TreeSpecies ;')
    lines.append(f'    rdfs:label "{common}"@en ;')
    lines.append(f'    ref:speciesCode "{code}" ;')
    if latin:
        lines.append(f'    ref:latinName "{latin}" ;')
    lines.append(f'    ref:mapsToModelGroup ref:modelGroup{model} .')
    lines.append('')

with open("docs/reference-data/tree-species.ttl", "w") as f:
    f.write('\n'.join(lines))
print(f"1. tree-species.ttl: {sum(1 for l in lines if 'a ref:TreeSpecies' in l)} species")

# ============================================================================
# 2. SPECIES MODEL GROUPS + YIELD CLASSES + SPACING OPTIONS (TTL)
# ============================================================================
ws_yc = wb["Species_YC_Ranges"]

# Row 2 has codes like "BE_SP", "CP_SP" in cols 6-24 (col 5 = "Select_SP")
# Spacings: rows 4-9, Yield classes: rows 13-24
model_groups = {}

for col in range(6, 25):
    sp_code = ws_yc.cell(2, col).value
    if not sp_code or str(sp_code).strip() == "":
        continue
    # Extract model code from "BE_SP" → "BE"
    code = str(sp_code).strip().replace("_SP", "")
    spacings = []
    yield_classes = []
    for row in range(4, 10):
        v = ws_yc.cell(row, col).value
        if v is not None and isinstance(v, (int, float)):
            spacings.append(float(v))
    for row in range(13, 25):
        v = ws_yc.cell(row, col).value
        if v is not None and isinstance(v, (int, float)):
            yield_classes.append(int(v))
    if spacings or yield_classes:
        model_groups[code] = {"spacings": spacings, "yieldClasses": yield_classes}

# Collect all unique spacings and yield classes
all_spacings = sorted(set(s for g in model_groups.values() for s in g["spacings"]))
all_yc = sorted(set(y for g in model_groups.values() for y in g["yieldClasses"]))

# Write yield-classes.ttl
yc_lines = [
    '@prefix rdfs:  <http://www.w3.org/2000/01/rdf-schema#> .',
    '@prefix xsd:   <http://www.w3.org/2001/XMLSchema#> .',
    '@prefix ref:   <https://w3id.org/wcc/reference#> .',
    '',
    '# =============================================================================',
    '# Reference Data — Yield Classes',
    '# =============================================================================',
    ''
]
for yc in all_yc:
    yc_lines.append(f'ref:yieldClass{yc}')
    yc_lines.append(f'    a ref:YieldClass ;')
    yc_lines.append(f'    rdfs:label "Yield Class {yc}"@en ;')
    yc_lines.append(f'    ref:yieldClassValue {yc} .')
    yc_lines.append('')

with open("docs/reference-data/yield-classes.ttl", "w") as f:
    f.write('\n'.join(yc_lines))
print(f"2. yield-classes.ttl: {len(all_yc)} yield classes: {all_yc}")

# ============================================================================
# 3. TREE SPACING OPTIONS with emission rates (TTL)
# ============================================================================
ws_vl = wb["Validation lists"]

# Validation lists rows 5-13: spacing, seedling, shelter1.2m, shelter0.6m(=spiral guard), voleguard, fertiliser
# Row 3 has per-tree rates, rows 5-13 have per-ha rates
sp_lines = [
    '@prefix rdfs:  <http://www.w3.org/2000/01/rdf-schema#> .',
    '@prefix xsd:   <http://www.w3.org/2001/XMLSchema#> .',
    '@prefix ref:   <https://w3id.org/wcc/reference#> .',
    '@prefix qty:   <https://w3id.org/wcc/quantity#> .',
    '',
    '# =============================================================================',
    '# Reference Data — Tree Spacing Options with Emission Rates',
    '# =============================================================================',
    '# Rates are in tCO2e per hectare (negative = emissions).',
    '# Shelter rates are for 1.2m tree shelters.',
    '# Spiral guard rates are for 0.6m spiral guards.',
    '# =============================================================================',
    ''
]

for row in range(5, 14):
    spacing = ws_vl.cell(row, 1).value
    if spacing is None:
        continue
    spacing = float(spacing)
    seedling = ws_vl.cell(row, 2).value or 0
    shelter_1_2 = ws_vl.cell(row, 3).value or 0
    spiral_0_6 = ws_vl.cell(row, 4).value or 0
    voleguard = ws_vl.cell(row, 5).value or 0
    fertiliser = ws_vl.cell(row, 6).value or 0

    sp_str = str(spacing).replace('.', 'p')
    sp_lines.append(f'ref:spacing{sp_str}')
    sp_lines.append(f'    a ref:TreeSpacingOption ;')
    sp_lines.append(f'    rdfs:label "{spacing}m spacing"@en ;')
    sp_lines.append(f'    ref:spacingValue {spacing} ;')
    sp_lines.append(f'    ref:seedlingRate {seedling} ;')
    sp_lines.append(f'    ref:voleguardRate {voleguard} ;')
    sp_lines.append(f'    ref:fertiliserRate {fertiliser} ;')
    # Shelter rate for 1.2m height
    sp_lines.append(f'    ref:hasShelterRate ref:shelter{sp_str}_1p2m ;')
    # Spiral guard rate for 0.6m height
    sp_lines.append(f'    ref:hasSpiralGuardRate ref:spiral{sp_str}_0p6m .')
    sp_lines.append('')
    # Shelter rate individual
    sp_lines.append(f'ref:shelter{sp_str}_1p2m')
    sp_lines.append(f'    a ref:ShelterRate ;')
    sp_lines.append(f'    rdfs:label "Shelter rate at {spacing}m spacing, 1.2m height"@en ;')
    sp_lines.append(f'    ref:protectionEmissionRate {shelter_1_2} .')
    sp_lines.append('')
    # Spiral guard rate individual
    sp_lines.append(f'ref:spiral{sp_str}_0p6m')
    sp_lines.append(f'    a ref:SpiralGuardRate ;')
    sp_lines.append(f'    rdfs:label "Spiral guard rate at {spacing}m spacing, 0.6m height"@en ;')
    sp_lines.append(f'    ref:protectionEmissionRate {spiral_0_6} .')
    sp_lines.append('')

with open("docs/reference-data/tree-spacing-options.ttl", "w") as f:
    f.write('\n'.join(sp_lines))
print(f"3. tree-spacing-options.ttl: {len(all_spacings)} spacing options")

# ============================================================================
# 4. SPECIES MODEL GROUPS (TTL)
# ============================================================================
mg_lines = [
    '@prefix rdfs:  <http://www.w3.org/2000/01/rdf-schema#> .',
    '@prefix xsd:   <http://www.w3.org/2001/XMLSchema#> .',
    '@prefix ref:   <https://w3id.org/wcc/reference#> .',
    '',
    '# =============================================================================',
    '# Reference Data — Species Model Groups',
    '# =============================================================================',
    '# Each model group has valid spacing options and yield classes.',
    '# =============================================================================',
    ''
]

for code, data in sorted(model_groups.items()):
    mg_lines.append(f'ref:modelGroup{code}')
    mg_lines.append(f'    a ref:SpeciesModelGroup ;')
    mg_lines.append(f'    rdfs:label "Model Group {code}"@en ;')
    mg_lines.append(f'    ref:speciesModelCode "{code}" ;')
    for sp in data["spacings"]:
        sp_str = str(sp).replace('.', 'p')
        mg_lines.append(f'    ref:hasAvailableSpacing ref:spacing{sp_str} ;')
    for yc in data["yieldClasses"]:
        mg_lines.append(f'    ref:hasAvailableYieldClass ref:yieldClass{yc} ;')
    # Fix trailing semicolon -> period
    mg_lines[-1] = mg_lines[-1].rstrip(' ;') + ' .'
    mg_lines.append('')

with open("docs/reference-data/species-model-groups.ttl", "w") as f:
    f.write('\n'.join(mg_lines))
print(f"4. species-model-groups.ttl: {len(model_groups)} groups: {sorted(model_groups.keys())}")

# ============================================================================
# 5. SOIL EMISSION LOOKUP (TTL)
# ============================================================================
# Validation lists rows 29-52: country+pct code, country, pct, seminatural, pasture, arable
soil_lines = [
    '@prefix rdfs:  <http://www.w3.org/2000/01/rdf-schema#> .',
    '@prefix xsd:   <http://www.w3.org/2001/XMLSchema#> .',
    '@prefix carb:  <https://w3id.org/wcc/carbon#> .',
    '@prefix ref:   <https://w3id.org/wcc/reference#> .',
    '',
    '# =============================================================================',
    '# Reference Data — Soil Carbon Emission Lookup',
    '# =============================================================================',
    '# Each entry provides soil carbon emissions (tCO2e/ha) for a specific',
    '# country and topsoil carbon loss percentage, broken down by previous landuse.',
    '# =============================================================================',
    ''
]

country_iri = {
    "England": "ref:England",
    "Scotland": "ref:Scotland",
    "Wales": "ref:Wales",
    "Northern Ireland": "ref:NorthernIreland"
}
count = 0
for row in range(29, 53):
    code = ws_vl.cell(row, 1).value
    country = ws_vl.cell(row, 2).value
    pct = ws_vl.cell(row, 3).value
    semi = ws_vl.cell(row, 4).value or 0
    past = ws_vl.cell(row, 5).value or 0
    arab = ws_vl.cell(row, 6).value or 0
    if not code or not country:
        continue
    code = str(code).strip()
    country = str(country).strip()
    pct = str(pct).strip()
    local = code.replace(' ', '')
    
    soil_lines.append(f'carb:soilEmission{local}')
    soil_lines.append(f'    a carb:SoilEmissionLookupEntry ;')
    soil_lines.append(f'    rdfs:label "Soil emission: {country}, {pct}% topsoil lost"@en ;')
    soil_lines.append(f'    carb:soilEmissionCountry {country_iri[country]} ;')
    soil_lines.append(f'    carb:topsoilPercentageCode "{pct}" ;')
    soil_lines.append(f'    carb:hasSoilEmissionRate carb:soilRate{local}Seminatural ;')
    soil_lines.append(f'    carb:hasSoilEmissionRate carb:soilRate{local}Pasture ;')
    soil_lines.append(f'    carb:hasSoilEmissionRate carb:soilRate{local}Arable .')
    soil_lines.append('')
    # Three rates per entry
    for landuse, val, iri_suffix in [("Seminatural", semi, "Seminatural"), ("Pasture", past, "Pasture"), ("Arable", arab, "Arable")]:
        soil_lines.append(f'carb:soilRate{local}{iri_suffix}')
        soil_lines.append(f'    a carb:SoilEmissionRate ;')
        soil_lines.append(f'    rdfs:label "Soil rate: {country}, {pct}%, {landuse}"@en ;')
        soil_lines.append(f'    carb:forPreviousLanduse ref:{iri_suffix} ;')
        soil_lines.append(f'    carb:soilEmissionValue {val} .')
        soil_lines.append('')
    count += 1

with open("docs/reference-data/soil-emission-lookup.ttl", "w") as f:
    f.write('\n'.join(soil_lines))
print(f"5. soil-emission-lookup.ttl: {count} entries (each with 3 landuse rates)")

# ============================================================================
# 6. SOIL CARBON ACCUMULATION RATES (TTL)
# ============================================================================
# Fixed values from column AY of the StandardProjectCarbonCalculator
# Read directly from the sheet
ws_calc = wb["StandardProjectCarbonCalculator"]
acc_lines = [
    '@prefix rdfs:  <http://www.w3.org/2000/01/rdf-schema#> .',
    '@prefix xsd:   <http://www.w3.org/2001/XMLSchema#> .',
    '@prefix carb:  <https://w3id.org/wcc/carbon#> .',
    '@prefix ref:   <https://w3id.org/wcc/reference#> .',
    '',
    '# =============================================================================',
    '# Reference Data — Soil Carbon Accumulation Rates',
    '# =============================================================================',
    '# Cumulative tCO2e/ha for previously-arable mineral soil under',
    '# minimum-intervention management, indexed by growth period.',
    '# =============================================================================',
    ''
]

# Column AY = column 51, rows correspond to periods
# Let's find which column AY is - AY is col 51
period_names = [
    ("period0to5", "0-5"),
    ("period5to10", "5-10"),
    ("period10to15", "10-15"),
    ("period15to20", "15-20"),
    ("period20to25", "20-25"),
    ("period25to30", "25-30"),
    ("period30to35", "30-35"),
    ("period35to40", "35-40"),
    ("period40to45", "40-45"),
    ("period45to50", "45-50"),
    ("period50to55", "50-55"),
    ("period55to60", "55-60"),
    ("period60to65", "60-65"),
    ("period65to70", "65-70"),
    ("period70to75", "70-75"),
    ("period75to80", "75-80"),
    ("period80to85", "80-85"),
    ("period85to90", "85-90"),
    ("period90to95", "90-95"),
    ("period95to100", "95-100"),
]

# Read AY column values from the sheet (rows 7-26 for periods 0-5 through 95-100)
for i, (period_id, period_label) in enumerate(period_names):
    row = 7 + i  # AY7 is first period (0-5)
    val = ws_calc.cell(row, 51).value  # AY = col 51
    if val is None:
        val = 0.0
    val = round(float(val), 4)
    acc_lines.append(f'carb:soilAccum{period_id.replace("period","").replace("to","_")}')
    acc_lines.append(f'    a carb:SoilCarbonAccumulationRate ;')
    acc_lines.append(f'    rdfs:label "Soil C accumulation: {period_label}"@en ;')
    acc_lines.append(f'    carb:accumulationForPeriod ref:{period_id} ;')
    acc_lines.append(f'    carb:accumulationValue {val} .')
    acc_lines.append('')

with open("docs/reference-data/soil-carbon-accumulation.ttl", "w") as f:
    f.write('\n'.join(acc_lines))
print(f"6. soil-carbon-accumulation.ttl: {len(period_names)} periods")

# ============================================================================
# 7. BIOMASS CARBON LOOKUP TABLE (JSON-LD) - ~19,000 rows
# ============================================================================
ws_bio = wb["Biomass carbon lookup table"]
# Col 1: lookup key, Col 2: Species, Col 3: Spacing, Col 4: YC, Col 5: Mgmt,
# Col 6: Period, Col 7-14: carbon values
print(f"\n7. Biomass lookup: {ws_bio.max_row - 1} data rows")

biomass_entries = []
for row in range(2, ws_bio.max_row + 1):
    lookup_key = ws_bio.cell(row, 1).value
    if lookup_key is None:
        continue
    species = ws_bio.cell(row, 2).value
    spacing = ws_bio.cell(row, 3).value
    yc = ws_bio.cell(row, 4).value
    mgmt = ws_bio.cell(row, 5).value
    period = ws_bio.cell(row, 6).value

    entry = {
        "lookupKey": str(lookup_key).strip(),
        "species": str(species).strip() if species else None,
        "spacing": float(spacing) if spacing is not None else None,
        "yieldClass": int(yc) if yc is not None else None,
        "managementRegime": str(mgmt).strip() if mgmt else None,
        "period": str(period).strip() if period else None,
    }
    # Col 7-14: carbon values
    col_keys = ["carbonStanding", "debrisCarbon", "totalCarbonRate",
                "cumulativeInPeriod", "cumulativeBiomassSeq",
                "cumulativeMgmtEmissions", "cumulativeTotalSeq",
                "removedFromForest"]
    for i, key in enumerate(col_keys):
        v = ws_bio.cell(row, 7 + i).value
        entry[key] = round(float(v), 6) if v is not None else 0.0
    
    biomass_entries.append(entry)

# Write as JSON-LD
biomass_jsonld = {
    "@context": {
        "carb": "https://w3id.org/wcc/carbon#",
        "ref": "https://w3id.org/wcc/reference#",
        "rdfs": "http://www.w3.org/2000/01/rdf-schema#",
        "xsd": "http://www.w3.org/2001/XMLSchema#",
        "lookupKey": "carb:biomassLookupKey",
        "species": "carb:speciesModelCode",
        "spacing": "carb:spacingValue",
        "yieldClass": "carb:yieldClassValue",
        "managementRegime": "carb:managementRegimeCode",
        "period": "carb:periodLabel",
        "carbonStanding": { "@id": "carb:carbonStandingValue", "@type": "xsd:double" },
        "debrisCarbon": { "@id": "carb:debrisCarbonValue", "@type": "xsd:double" },
        "totalCarbonRate": { "@id": "carb:totalCarbonRateValue", "@type": "xsd:double" },
        "cumulativeInPeriod": { "@id": "carb:cumulativeInPeriodValue", "@type": "xsd:double" },
        "cumulativeBiomassSeq": { "@id": "carb:cumulativeBiomassSeqValue", "@type": "xsd:double" },
        "removedFromForest": { "@id": "carb:removedFromForestValue", "@type": "xsd:double" },
        "cumulativeMgmtEmissions": { "@id": "carb:cumulativeMgmtEmissionsValue", "@type": "xsd:double" },
        "cumulativeTotalSeq": { "@id": "carb:cumulativeTotalSeqValue", "@type": "xsd:double" }
    },
    "@graph": biomass_entries
}

with open("docs/reference-data/biomass-carbon-lookup.jsonld", "w") as f:
    json.dump(biomass_jsonld, f)  # No indent to keep file size reasonable
print(f"7. biomass-carbon-lookup.jsonld: {len(biomass_entries)} entries")

# ============================================================================
# 8. CLEARFELL MAX SEQUESTRATION (JSON-LD) - ~26,000 rows
# ============================================================================
ws_cf = wb["Clearfell max seq values"]
# Row 3 has headers: col 2=Species, col 3=Spacing, col 4=YC, col 5=Mgmt, cols 6-37=rotation ages
# Row 1 col 1 has lookup key, data starts at row 4

# Get rotation ages from row 3, cols 6+
clearfell_ages = []
for c in range(6, 38):
    h = ws_cf.cell(3, c).value
    if h is not None and isinstance(h, (int, float)):
        clearfell_ages.append(int(h))

print(f"\n8. Clearfell: rotation ages = {clearfell_ages}")

clearfell_entries = []
for row in range(4, ws_cf.max_row + 1):
    lookup_key = ws_cf.cell(row, 1).value
    if lookup_key is None:
        continue
    species = ws_cf.cell(row, 2).value
    spacing = ws_cf.cell(row, 3).value
    yc = ws_cf.cell(row, 4).value
    mgmt = ws_cf.cell(row, 5).value
    
    caps = {}
    for i, age in enumerate(clearfell_ages):
        v = ws_cf.cell(row, 6 + i).value
        if v is not None:
            caps[str(age)] = round(float(v), 6)
    
    clearfell_entries.append({
        "lookupKey": str(lookup_key).strip(),
        "species": str(species).strip() if species else None,
        "spacing": float(spacing) if spacing is not None else None,
        "yieldClass": int(yc) if yc is not None else None,
        "managementRegime": str(mgmt).strip() if mgmt else None,
        "maxSequestrationByAge": caps
    })

clearfell_jsonld = {
    "@context": {
        "carb": "https://w3id.org/wcc/carbon#",
        "ref": "https://w3id.org/wcc/reference#",
        "xsd": "http://www.w3.org/2001/XMLSchema#",
        "species": "carb:speciesModelCode",
        "spacing": "carb:spacingValue",
        "yieldClass": "carb:yieldClassValue",
        "managementRegime": "carb:managementRegimeCode",
        "lookupKey": "carb:clearfellLookupKey",
        "maxSequestrationByAge": "carb:maxSequestrationByAge"
    },
    "@graph": clearfell_entries
}

with open("docs/reference-data/clearfell-max-seq.jsonld", "w") as f:
    json.dump(clearfell_jsonld, f)
print(f"8. clearfell-max-seq.jsonld: {len(clearfell_entries)} entries, ages: {clearfell_ages}")

# ============================================================================
# 9. Topsoil carbon % lost by ground prep + soil type (supplement for TTL)
# ============================================================================
# Already in ground-preparation-methods.ttl but only single code.
# The Validation lists rows 17-23 show the % differs by soil type.
# Let's update the ground prep topsoil codes to include both mineral and organomineral.
print("\n9. Ground prep topsoil % by soil type:")
for row in range(17, 24):
    method = ws_vl.cell(row, 1).value
    organo = ws_vl.cell(row, 2).value  
    mineral = ws_vl.cell(row, 3).value
    print(f"   {method}: organomineral={organo}, mineral={mineral}")

print("\nDone! All files written.")
