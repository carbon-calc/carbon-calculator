import openpyxl
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('docs/CarbonCalculator_Examples_V3.0_August2025.xlsx', data_only=True)
ws = wb['Natural regeneration']

# Part 1: First 7 rows, all columns
print("=== ROWS 1-7 ===")
for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=52):
    for cell in row:
        if cell.value is not None:
            print(f'{cell.coordinate}: {repr(cell.value)}')

# Part 2: The planting section details - H columns area
print("\n=== PLANTING SECTION H-W, rows 4-7 ===")
for row in ws.iter_rows(min_row=1, max_row=7, min_col=8, max_col=23):
    for cell in row:
        if cell.value is not None:
            print(f'{cell.coordinate}: {repr(cell.value)}')

# Part 3: Results table - rows 52-100
print("\n=== ROWS 52-100 ===")
for row in ws.iter_rows(min_row=52, max_row=100, min_col=1, max_col=52):
    for cell in row:
        if cell.value is not None:
            print(f'{cell.coordinate}: {repr(cell.value)}')
