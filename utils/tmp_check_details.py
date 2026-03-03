import openpyxl
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook('docs/CarbonCalculator_Examples_V3.0_August2025.xlsx', data_only=True)

# Check establishment emission details for Broadleaved sheet (has -556 total)
ws = wb['Broadleaved min intervention']
print("=== Broadleaved min intervention: Establishment emissions ===")
for row_num in range(16, 35):
    for col in [1, 2, 3, 4, 5]:  # A-E
        v = ws.cell(row=row_num, column=col).value
        if v is not None:
            cell_ref = ws.cell(row=row_num, column=col).coordinate
            print(f"  {cell_ref}: {repr(v)}")

# Check planting sections for all standard sheets
print("\n=== Planting section details for all 4 standard sheets ===")
for name in ['Natural regeneration', 'Broadleaved min intervention', 'Mixed conifer thin', 'Conifer clearfell']:
    ws = wb[name]
    print(f"\n--- {name} ---")
    for row_num in range(6, 33):
        h_val = ws.cell(row=row_num, column=8).value  # H
        g_val = ws.cell(row=row_num, column=7).value  # G
        if h_val and 'Select' not in str(h_val):
            s_val = ws.cell(row=row_num, column=19).value  # S
            t_val = ws.cell(row=row_num, column=20).value  # T
            u_val = ws.cell(row=row_num, column=21).value  # U
            r_val = ws.cell(row=row_num, column=18).value  # R
            i_val = ws.cell(row=row_num, column=9).value   # I
            p_val = ws.cell(row=row_num, column=16).value  # P
            print(f"  Row {row_num}: G={g_val} H={h_val} I={i_val} P(YC)={p_val} R={r_val} S={s_val} T={t_val} U={u_val}")
