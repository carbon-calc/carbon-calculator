import openpyxl, sys, warnings
warnings.filterwarnings("ignore")
wb = openpyxl.load_workbook("docs/CarbonCalculator_V3.0_August2025.xlsx", data_only=True)

print("=== ALL SHEETS ===")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row}r x {ws.max_column}c")

print("\n=== ALL NAMED RANGES ===")
for name, dn in wb.defined_names.items():
    print(f"  {name} = {dn.attr_text}")

# Explore Validation lists for spacing emission rates
ws = wb["Validation lists"]
print(f"\n=== Validation lists: rows 1-25, cols 1-12 ===")
for r in range(1, 26):
    vals = []
    for c in range(1, 13):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append(f"[{c}]={v}")
    if vals:
        print(f"  R{r}: {', '.join(vals)}")

# Soil carbon emission table area
print(f"\n=== Validation lists: rows 25-80, cols 1-12 ===")
for r in range(25, min(81, ws.max_row+1)):
    vals = []
    for c in range(1, 13):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append(f"[{c}]={v}")
    if vals:
        print(f"  R{r}: {', '.join(vals)}")
